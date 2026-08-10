 """
Obesity Expert System — Core Engine
=====================================
Implements the confirmed decisions from the requirements/decision-rules document:

1. BMI classification (6 categories)
2. BMR (Mifflin-St Jeor) -> TDEE
3. Calorie deficit: Fuzzy Inference System (Mamdani) over BMI + Activity
   -> deficit %, then Lower/Upper calorie targets
4. Minimum calorie floor: Male = 1500 kcal / Female = 1200 kcal
5. Macronutrient split: classified into one of Asmaa's 3 reference scenarios
   (Balanced 50/25/25, High-Protein 45/30/25, High-Carb 60/20/20) via a
   simple fuzzy rule set on BMI + Activity, instead of a continuous %.
6. Portion sizes: looked up directly from Asmaa's pre-built reference table
   (PORTION_SIZE_EXCEL.xlsx) at the nearest matching calorie level + scenario,
   so no food-exchange math is repeated from scratch for every client.

This is a first working version meant to be tested against real client cases
and refined, not a final production system.
"""

import numpy as np
import skfuzzy as fuzz
from skfuzzy import control as ctrl
from openpyxl import load_workbook
from dataclasses import dataclass


# ---------------------------------------------------------------------------
# 1. Basic clinical calculations
# ---------------------------------------------------------------------------

def calculate_bmi(weight_kg: float, height_cm: float) -> float:
    return weight_kg / ((height_cm / 100) ** 2)


def classify_bmi(bmi: float) -> str:
    if bmi < 18.5:
        return "Underweight"
    elif bmi < 25:
        return "Normal weight"
    elif bmi < 30:
        return "Overweight"
    elif bmi < 35:
        return "Obesity I"
    elif bmi < 40:
        return "Obesity II"
    else:
        return "Severe obesity"


def calculate_bmr(weight_kg: float, height_cm: float, age: int, sex: str) -> float:
    """Mifflin-St Jeor equation. sex: 'M' or 'F'"""
    base = 10 * weight_kg + 6.25 * height_cm - 5 * age
    if sex.upper() == "M":
        return base + 5
    elif sex.upper() == "F":
        return base - 161
    else:
        raise ValueError("sex must be 'M' or 'F'")


def calculate_tdee(bmr: float, activity_factor: float) -> float:
    return bmr * activity_factor


def calorie_floor(sex: str) -> int:
    """Confirmed decision: Male = 1500 kcal, Female = 1200 kcal."""
    return 1500 if sex.upper() == "M" else 1200


# ---------------------------------------------------------------------------
# 2. Fuzzy Inference System — Calorie Deficit % (BMI + Activity -> Deficit%)
# ---------------------------------------------------------------------------

def _build_deficit_fis():
    bmi = ctrl.Antecedent(np.arange(15, 50.1, 0.1), 'bmi')
    activity = ctrl.Antecedent(np.arange(1.2, 1.9001, 0.001), 'activity')
    deficit = ctrl.Consequent(np.arange(0, 30.1, 0.1), 'deficit')

    bmi['normal'] = fuzz.trapmf(bmi.universe, [15, 15, 23, 25])
    bmi['overweight'] = fuzz.trapmf(bmi.universe, [23, 25, 28, 30])
    bmi['obesity1'] = fuzz.trapmf(bmi.universe, [28, 30, 33, 35])
    bmi['obesity2'] = fuzz.trapmf(bmi.universe, [33, 35, 38, 40])
    bmi['severe'] = fuzz.trapmf(bmi.universe, [38, 40, 50, 50])

    activity['sedentary'] = fuzz.trapmf(activity.universe, [1.2, 1.2, 1.2, 1.3])
    activity['light'] = fuzz.trimf(activity.universe, [1.2, 1.375, 1.55])
    activity['moderate'] = fuzz.trimf(activity.universe, [1.375, 1.55, 1.725])
    activity['very'] = fuzz.trimf(activity.universe, [1.55, 1.725, 1.9])
    activity['extreme'] = fuzz.trapmf(activity.universe, [1.725, 1.9, 1.9, 1.9])

    deficit['none'] = fuzz.trimf(deficit.universe, [0, 0, 8])
    deficit['mild'] = fuzz.trimf(deficit.universe, [5, 12, 18])
    deficit['moderate'] = fuzz.trimf(deficit.universe, [15, 20, 25])
    deficit['high'] = fuzz.trimf(deficit.universe, [20, 25, 30])
    deficit['very_high'] = fuzz.trimf(deficit.universe, [25, 30, 30])

    rules = [
        ctrl.Rule(bmi['normal'], deficit['none']),
        ctrl.Rule(bmi['overweight'], deficit['mild']),
        ctrl.Rule(bmi['obesity1'], deficit['mild']),
        ctrl.Rule(bmi['obesity2'], deficit['moderate']),
        ctrl.Rule(bmi['severe'] & activity['sedentary'], deficit['very_high']),
        ctrl.Rule(bmi['severe'] & (activity['light'] | activity['moderate']), deficit['high']),
        ctrl.Rule(bmi['severe'] & (activity['very'] | activity['extreme']), deficit['moderate']),
        ctrl.Rule(activity['very'] | activity['extreme'], deficit['mild']),
    ]
    system = ctrl.ControlSystem(rules)
    return system


_deficit_system = _build_deficit_fis()


def fuzzy_deficit_percent(bmi: float, activity_factor: float) -> float:
    sim = ctrl.ControlSystemSimulation(_deficit_system)
    sim.input['bmi'] = bmi
    sim.input['activity'] = activity_factor
    sim.compute()
    return sim.output['deficit']


# ---------------------------------------------------------------------------
# 3. Macro scenario classification (3 discrete scenarios, fuzzy-selected)
# ---------------------------------------------------------------------------

MACRO_SCENARIOS = {
    "balanced": {"carb": 0.50, "protein": 0.25, "fat": 0.25,
                 "label": "Normal macronutrients distribution (50% Carb,25% Protein,Fat )"},
    "high_protein": {"carb": 0.45, "protein": 0.30, "fat": 0.25,
                      "label": "Normal macronutrients distribution (45% Carb,30% Protein,25% Fat )"},
    "high_carb": {"carb": 0.60, "protein": 0.20, "fat": 0.20,
                   "label": "Normal macronutrients distribution (60% Carb,20% Protein,20% Fat )"},
    "moderate_carb": {"carb": 0.55, "protein": 0.20, "fat": 0.25,
                        "label": "Normal macronutrients distribution (55% Carb,20% Protein,25% Fat )"},
}


def classify_macro_scenario(bmi: float, activity_factor: float) -> str:
    """
    DEPRECATED for direct use — kept for backward compatibility / quick checks.
    Returns only the single strongest-firing scenario. For real client
    decisions, use get_macro_scenario_candidates() instead, since Asmaa's
    confirmed decision is that when multiple scenarios are plausible at
    once, the specialist should see all of them and choose — not have the
    system silently pick one.
    """
    candidates = get_macro_scenario_candidates(bmi, activity_factor)
    return candidates[0][0] if candidates else "balanced"


def get_macro_scenario_candidates(bmi: float, activity_factor: float, threshold: float = 0.3):
    """
    Returns every macro scenario whose fuzzy membership/firing strength is
    above `threshold`, sorted strongest-first, instead of collapsing to a
    single crisp choice.

    Confirmed decision (Asmaa): when two or more scenarios are plausible
    at the same time, the system should present the options and let the
    specialist decide — this applies especially to macro-scenario
    selection and the Lower/Upper calorie target (which already always
    returns both, unchanged).

    Returns: list of (scenario_key, strength) tuples, strength in [0,1].
    """
    bmi_u = np.arange(15, 50.1, 0.1)
    act_u = np.arange(1.2, 1.9001, 0.001)

    mf_normal = fuzz.trapmf(bmi_u, [15, 15, 23, 25])
    mf_overweight = fuzz.trapmf(bmi_u, [23, 25, 28, 30])
    mf_obesity1 = fuzz.trapmf(bmi_u, [28, 30, 33, 35])
    mf_obesity2 = fuzz.trapmf(bmi_u, [33, 35, 38, 40])
    mf_severe = fuzz.trapmf(bmi_u, [38, 40, 50, 50])

    mf_sedentary = fuzz.trapmf(act_u, [1.2, 1.2, 1.2, 1.3])
    mf_light = fuzz.trimf(act_u, [1.2, 1.375, 1.55])
    mf_moderate = fuzz.trimf(act_u, [1.375, 1.55, 1.725])
    mf_very = fuzz.trimf(act_u, [1.55, 1.725, 1.9])
    mf_extreme = fuzz.trapmf(act_u, [1.725, 1.9, 1.9, 1.9])

    d = lambda mf, u, x: fuzz.interp_membership(u, mf, x)

    normal = d(mf_normal, bmi_u, bmi)
    overweight = d(mf_overweight, bmi_u, bmi)
    obesity1 = d(mf_obesity1, bmi_u, bmi)
    obesity2 = d(mf_obesity2, bmi_u, bmi)
    severe = d(mf_severe, bmi_u, bmi)

    sedentary = d(mf_sedentary, act_u, activity_factor)
    light = d(mf_light, act_u, activity_factor)
    moderate = d(mf_moderate, act_u, activity_factor)
    very = d(mf_very, act_u, activity_factor)
    extreme = d(mf_extreme, act_u, activity_factor)

    firing = {
        "high_carb": max(very, extreme),
        "moderate_carb": moderate,
        "high_protein": min(max(obesity2, severe), max(sedentary, light)),
        "balanced": max(normal, overweight, obesity1),
    }

    candidates = [(k, round(v, 2)) for k, v in firing.items() if v >= threshold]
    candidates.sort(key=lambda kv: kv[1], reverse=True)
    return candidates if candidates else [("balanced", 0.0)]


# ---------------------------------------------------------------------------
# 4. Portion lookup from Asmaa's pre-built reference table
# ---------------------------------------------------------------------------

PORTION_COLUMNS = [
    "free_fat_milk", "low_fat_milk", "whole_milk", "cooked_vegetable",
    "green_salad", "fruit", "starch", "very_lean_meat", "lean_meat",
    "medium_fat_meat", "high_fat_meat", "fats_and_oil",
]


def load_portion_reference(path: str):
    """Parses PORTION_SIZE_EXCEL.xlsx into a list of reference rows."""
    wb = load_workbook(path, data_only=True)
    ws = wb["Sheet1"]
    rows = []
    current_level = None
    for r in range(1, ws.max_row + 1):
        level_cell = ws.cell(row=r, column=1).value
        label_cell = ws.cell(row=r, column=2).value
        if isinstance(level_cell, (int, float)):
            current_level = level_cell
        if label_cell and "Normal macronutrients distribution" in str(label_cell):
            portions = {}
            for i, col_letter_idx in enumerate(range(7, 19)):  # G..R
                val = ws.cell(row=r, column=col_letter_idx).value
                portions[PORTION_COLUMNS[i]] = val or 0
            actual_cal = ws.cell(row=r, column=19).value  # S
            rows.append({
                "calorie_level": current_level,
                "scenario_label": label_cell,
                "portions": portions,
                "actual_calorie": actual_cal,
            })
    return rows


def _scenario_key_from_label(label: str) -> str:
    if "50% Carb" in label:
        return "balanced"
    if "45% Carb" in label:
        return "high_protein"
    if "60% Carb" in label:
        return "high_carb"
    if "55% Carb" in label:
        return "moderate_carb"
    return "unknown"


def lookup_portions(reference_rows, target_calorie: float, scenario: str):
    """Finds the closest pre-computed row for the given scenario."""
    candidates = [r for r in reference_rows if _scenario_key_from_label(r["scenario_label"]) == scenario]
    if not candidates:
        return None
    best = min(candidates, key=lambda r: abs(r["calorie_level"] - target_calorie))
    return best


# ---------------------------------------------------------------------------
# 5. End-to-end pipeline
# ---------------------------------------------------------------------------

@dataclass
class ClientResult:
    bmi: float
    bmi_class: str
    bmr: float
    tdee: float
    deficit_percent: float
    lower_target: float
    upper_target: float
    macro_scenario_candidates: list       # e.g. [("balanced", 0.8), ("moderate_carb", 0.4)]
    portions_by_scenario: dict            # scenario_key -> portions dict, for each candidate


def run_pipeline(age, height_cm, weight_kg, sex, activity_factor, reference_rows):
    bmi = calculate_bmi(weight_kg, height_cm)
    bmi_class = classify_bmi(bmi)
    bmr = calculate_bmr(weight_kg, height_cm, age, sex)
    tdee = calculate_tdee(bmr, activity_factor)

    deficit_pct = fuzzy_deficit_percent(bmi, activity_factor)
    floor = calorie_floor(sex)

    if bmi < 25:
        lower_target = upper_target = tdee
    else:
        lower_target = max(tdee - (deficit_pct * 0.6 / 100) * tdee, floor)  # smaller cut
        upper_target = max(tdee - (deficit_pct / 100) * tdee, floor)        # full fuzzy cut

    # Confirmed decision: don't collapse to one scenario — show every
    # plausible candidate and let the specialist pick.
    candidates = get_macro_scenario_candidates(bmi, activity_factor)

    portions_by_scenario = {}
    for scenario_key, _strength in candidates:
        ref_row = lookup_portions(reference_rows, upper_target, scenario_key)
        portions_by_scenario[scenario_key] = ref_row["portions"] if ref_row else {}

    return ClientResult(
        bmi=round(bmi, 1), bmi_class=bmi_class, bmr=round(bmr),
        tdee=round(tdee), deficit_percent=round(deficit_pct, 1),
        lower_target=round(lower_target), upper_target=round(upper_target),
        macro_scenario_candidates=candidates,
        portions_by_scenario=portions_by_scenario,
    )


if __name__ == "__main__":
    ref = load_portion_reference("/mnt/user-data/uploads/PORTION_SIZE_EXCEL.xlsx")
    print(f"Loaded {len(ref)} reference rows from Asmaa's portion table.\n")

    # Sample synthetic test cases (not real client data)
    test_cases = [
        dict(age=32, height_cm=166, weight_kg=95, sex="F", activity_factor=1.2),
        dict(age=45, height_cm=178, weight_kg=110, sex="M", activity_factor=1.375),
        dict(age=28, height_cm=170, weight_kg=80, sex="F", activity_factor=1.725),
    ]

    for i, case in enumerate(test_cases, 1):
        result = run_pipeline(**case, reference_rows=ref)
        print(f"--- Test case {i} ({case['sex']}, age {case['age']}) ---")
        print(f"BMI: {result.bmi} ({result.bmi_class})")
        print(f"BMR: {result.bmr} kcal | TDEE: {result.tdee} kcal")
        print(f"Fuzzy deficit: {result.deficit_percent}%")
        print(f"Calorie range: {result.lower_target}-{result.upper_target} kcal")
        print(f"Macro scenario candidates (for specialist to choose from):")
        for scenario_key, strength in result.macro_scenario_candidates:
            label = MACRO_SCENARIOS[scenario_key]["label"]
            print(f"  - {scenario_key} (match: {strength}): {label}")
            print(f"    portions: {result.portions_by_scenario[scenario_key]}")
        print()